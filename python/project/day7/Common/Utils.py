# 读取excel文件
import openpyxl
from day5.生成10个学生的成绩 import StudentInfo 
class ExcelOperator:
    def __init__(self,path:str,sheet:str,infos,flag:bool=True):
        self.path = path
        self.sheet = sheet
        self.infos = infos #如果格式需要dict类型，需要提供key,infos就是提供的key
        self.flag = flag #如果flag是true，说明有表头，意味着数据从第二行开始；反之为false，没有表头
        #保存excel数据的集合
        self.student01 = []   #存储的文件格式为[[],[],[].....]
        self.student02 = [] #存储的文件格式为[{},{},{}.....]

    def read_file(self):
        #实例化一个
        workbook = openpyxl.load_workbook(self.path)
        # 定义一个sheet
        sheet = workbook[self.sheet]
        # 开始读取
        for index,row in enumerate(sheet.rows):   #sheet.rows()返回这个表格有多少行
            # 判断有没有表头，如果有表头就跳出当前循环
            if self.flag and index == 0:
                continue
            # 定义一个集合存储,格式为[[],[],[].....]
            one_row_list = []
             # 定义一个集合存储,格式为[{},{},{}.....]
            one_row_dict = {}
            for col_index,col_value in enumerate(row):
                #把每一个单元格附加到one_row_list中
                one_row_list.append(col_value.value)
                one_row_dict[self.infos[col_index]] = col_value.value

            # 附加到总的集合中,格式为[[],[],[].....]
            self.student01.append(one_row_list)
             # 附加到总的集合中,格式为[{},{},{}.....]
            self.student02.append(one_row_dict)

    
 
class ExcelWrite:
    def __init__(self,path:str,sheet:str):
        self.path = path
        self.sheet = sheet
    def write_file(self,data):
        #  实例化一个workbook
        workbook = openpyxl.Workbook()
        # 激活一个sheet
        sheet = workbook.active
        # 为sheet设置一个title
        sheet.title = self.sheet
        # 开始遍历,[[],[],[].....]
        for index,row in enumerate(data):
            # 遍历每一行
            for col_index,col_val in enumerate(row):
                # 写入
                sheet.cell(row = index +1, column = col_index +1, value = col_val)
        #  写入真是的excel
        workbook.save(self.path)
        print('写入成功!')



if __name__ == '__main__':

    """==================读取===================
    # 准备一个path
    path = r'D:\Sandra\course\python\project\day7\student.xlsx'
    # 准备一个sheet
    sheet = 'Sheet2'
    infos = ['姓名','学号','年龄','性别','分数']
    # 实例化一个对象
    obj01 = ExcelOperator(path,sheet,infos,True)
    # 读取
    obj01.read_file()
    # 打印
    print(obj01.student01)
    print("\n")
    print(obj01.student02)
    """

    # 写入
    path=r'D:\Sandra\course\python\project\day7\Exam.xlsx'
    sheet = '初三3班'
    #准备数据
    obj01 = StudentInfo()
    # 生成数据
    obj01.get_one()
    # 实例化
    obj_write = ExcelWrite(path,sheet)
    # 写入
    obj_write.write_file(obj.results01)

